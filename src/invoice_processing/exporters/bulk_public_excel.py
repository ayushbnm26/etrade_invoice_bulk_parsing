from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from invoice_processing.bulk_models import PerFileOutcome
from invoice_processing.error_handling import PUBLIC_ACTION_NOTE
from invoice_processing.exceptions import ExportError
from invoice_processing.exporters.public_transform import (
    PUBLIC_AMOUNT_FIELDS,
    PUBLIC_COLUMN_LABELS,
    PUBLIC_ITEM_COLUMNS,
    PUBLIC_RATE_FIELDS,
    public_item_record,
)


SUMMARY_COLUMNS = [
    "Source File",
    "Sheet Name",
    "Invoice Number",
    "System Reference Number",
    "Status",
    "Item Rows",
    "Error Type",
    "Error Message",
]


class BulkPublicWorkbookExporter:
    def export(self, outcomes: list[PerFileOutcome], output_path: Path) -> Path:
        output_path.parent.mkdir(parents=True, exist_ok=True)

        try:
            workbook = Workbook()
            summary_sheet = workbook.active
            summary_sheet.title = "Processing Summary"
            self._build_summary_sheet(summary_sheet, outcomes)

            for outcome in outcomes:
                worksheet = workbook.create_sheet(outcome.sheet_name)
                if outcome.succeeded:
                    self._build_success_sheet(worksheet, outcome)
                else:
                    self._build_failure_sheet(worksheet, outcome)

            workbook.save(output_path)
        except Exception as exc:
            raise ExportError(f"Unable to write team workbook '{output_path.name}': {exc}") from exc

        return output_path

    def _build_summary_sheet(self, worksheet: Any, outcomes: list[PerFileOutcome]) -> None:
        self._prepare_sheet(worksheet)
        border = self._thin_border()
        header_fill = PatternFill("solid", fgColor="DDEBF2")
        text_color = "1F2937"

        for column_index, label in enumerate(SUMMARY_COLUMNS, start=1):
            cell = worksheet.cell(row=1, column=column_index, value=label)
            cell.fill = header_fill
            cell.font = Font(bold=True, color=text_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row_index, outcome in enumerate(outcomes, start=2):
            error = outcome.error
            values = [
                outcome.upload.filename,
                outcome.sheet_name,
                outcome.invoice_number,
                outcome.system_ref_no,
                outcome.status,
                outcome.item_row_count,
                error.error_type if error else "",
                error.message if error else "",
            ]
            for column_index, value in enumerate(values, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=value)
                cell.font = Font(color=text_color)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        self._auto_fit_columns(worksheet, max_width=36)

    def _build_success_sheet(self, worksheet: Any, outcome: PerFileOutcome) -> None:
        self._prepare_sheet(worksheet)
        border = self._thin_border()
        title_fill = PatternFill("solid", fgColor="EAF2F8")
        header_fill = PatternFill("solid", fgColor="DDEBF2")
        text_color = "1F2937"

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(PUBLIC_ITEM_COLUMNS))
        title = worksheet.cell(row=1, column=1, value="Invoice Items")
        title.font = Font(size=16, bold=True, color=text_color)
        title.fill = title_fill
        title.alignment = Alignment(horizontal="center", vertical="center")
        title.border = border
        worksheet.row_dimensions[1].height = 30

        for column_index in range(1, len(PUBLIC_ITEM_COLUMNS) + 1):
            worksheet.cell(row=1, column=column_index).border = border

        header_row = 3
        for column_index, label in enumerate(PUBLIC_COLUMN_LABELS, start=1):
            cell = worksheet.cell(row=header_row, column=column_index, value=label)
            cell.fill = header_fill
            cell.font = Font(bold=True, color=text_color)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        worksheet.row_dimensions[header_row].height = 36

        assert outcome.parsed_invoice is not None
        header = outcome.parsed_invoice.header
        public_items = [public_item_record(item, header) for item in outcome.parsed_invoice.line_items]

        for row_index, item in enumerate(public_items, start=header_row + 1):
            for column_index, (_, field_name) in enumerate(PUBLIC_ITEM_COLUMNS, start=1):
                cell = worksheet.cell(row=row_index, column=column_index, value=item.get(field_name, ""))
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.font = Font(color=text_color)
                self._apply_number_format(cell, field_name)

        max_row = max(header_row, header_row + len(public_items))
        self._append_address_footer(
            worksheet,
            header,
            start_row=max_row + 2,
            border=border,
            text_color=text_color,
        )
        worksheet.freeze_panes = "A4"
        worksheet.auto_filter.ref = f"A3:{get_column_letter(len(PUBLIC_ITEM_COLUMNS))}{max_row}"
        self._set_success_widths(worksheet)

    def _build_failure_sheet(self, worksheet: Any, outcome: PerFileOutcome) -> None:
        self._prepare_sheet(worksheet)
        border = self._thin_border()
        title_fill = PatternFill("solid", fgColor="F2E8E8")
        label_fill = PatternFill("solid", fgColor="F7F9FB")
        text_color = "1F2937"

        worksheet.merge_cells("A1:D1")
        title = worksheet["A1"]
        title.value = "Invoice Processing Failed"
        title.font = Font(size=16, bold=True, color=text_color)
        title.fill = title_fill
        title.alignment = Alignment(horizontal="center", vertical="center")
        title.border = border
        worksheet.row_dimensions[1].height = 30

        error = outcome.error
        rows = [
            ("Source Filename", outcome.upload.filename),
            ("Status", "FAILED"),
            ("Safe Error Type", error.error_type if error else "UnexpectedProcessingError"),
            ("Safe Error Message", error.message if error else "Processing failed."),
            ("Likely Reason", error.likely_reason if error else "unexpected processing failure"),
            ("Action Required", PUBLIC_ACTION_NOTE),
        ]

        for row_index, (label, value) in enumerate(rows, start=3):
            label_cell = worksheet.cell(row=row_index, column=1, value=label)
            value_cell = worksheet.cell(row=row_index, column=2, value=value)
            label_cell.fill = label_fill
            label_cell.font = Font(bold=True, color=text_color)
            value_cell.font = Font(color=text_color)
            for cell in (label_cell, value_cell):
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        worksheet.column_dimensions["A"].width = 24
        worksheet.column_dimensions["B"].width = 72
        worksheet.column_dimensions["C"].width = 12
        worksheet.column_dimensions["D"].width = 12

    def _prepare_sheet(self, worksheet: Any) -> None:
        worksheet.sheet_view.showGridLines = False

    def _append_address_footer(
        self,
        worksheet: Any,
        header: dict[str, Any],
        *,
        start_row: int,
        border: Border,
        text_color: str,
    ) -> None:
        heading_fill = PatternFill("solid", fgColor="EAF2F8")
        body_fill = PatternFill("solid", fgColor="FBFCFE")
        body_start_row = start_row + 1
        body_end_row = start_row + 5
        address_blocks = [
            (1, 8, "Shipping Address", self._format_address(header, "shipping")),
            (
                9,
                len(PUBLIC_ITEM_COLUMNS),
                "Receiver Shipping Address",
                self._format_address(header, "receiver_shipping"),
            ),
        ]

        worksheet.row_dimensions[start_row].height = 22
        for row_index in range(body_start_row, body_end_row + 1):
            worksheet.row_dimensions[row_index].height = 20

        for start_column, end_column, label, address in address_blocks:
            worksheet.merge_cells(
                start_row=start_row,
                start_column=start_column,
                end_row=start_row,
                end_column=end_column,
            )
            heading_cell = worksheet.cell(row=start_row, column=start_column, value=label)
            heading_cell.fill = heading_fill
            heading_cell.font = Font(bold=True, color=text_color)
            heading_cell.alignment = Alignment(horizontal="center", vertical="center")

            worksheet.merge_cells(
                start_row=body_start_row,
                start_column=start_column,
                end_row=body_end_row,
                end_column=end_column,
            )
            body_cell = worksheet.cell(
                row=body_start_row,
                column=start_column,
                value=address or "Not available",
            )
            body_cell.fill = body_fill
            body_cell.font = Font(color=text_color)
            body_cell.alignment = Alignment(vertical="top", wrap_text=True)

            for row_cells in worksheet.iter_rows(
                min_row=start_row,
                max_row=body_end_row,
                min_col=start_column,
                max_col=end_column,
            ):
                for cell in row_cells:
                    cell.border = border
                    if cell.row > start_row:
                        cell.fill = body_fill

    def _format_address(self, header: dict[str, Any], prefix: str) -> str:
        lines = [
            header.get(f"{prefix}_name", ""),
            header.get(f"{prefix}_address", ""),
        ]

        state_code = header.get(f"{prefix}_state_code", "")
        gstin = header.get(f"{prefix}_gstin", "")
        pan = header.get(f"{prefix}_pan", "")

        if state_code:
            lines.append(f"State Code: {state_code}")
        if gstin:
            lines.append(f"GSTIN: {gstin}")
        if pan:
            lines.append(f"PAN: {pan}")

        return "\n".join(str(line) for line in lines if line)

    def _apply_number_format(self, cell: Any, field_name: str) -> None:
        if field_name == "qty":
            cell.number_format = "#,##0"
        elif field_name in PUBLIC_RATE_FIELDS:
            cell.number_format = "0.00"
        elif field_name in PUBLIC_AMOUNT_FIELDS:
            cell.number_format = "#,##0.00"

    def _set_success_widths(self, worksheet: Any) -> None:
        preferred = {
            "A": 18,
            "B": 22,
            "C": 16,
            "D": 12,
            "E": 14,
            "F": 14,
            "G": 12,
            "H": 14,
            "I": 12,
            "J": 14,
            "K": 12,
            "L": 14,
            "M": 16,
            "N": 14,
            "O": 16,
            "P": 14,
        }
        self._auto_fit_columns(worksheet, max_width=26, min_width=11, preferred=preferred)

    def _auto_fit_columns(
        self,
        worksheet: Any,
        *,
        max_width: int = 60,
        min_width: int = 10,
        preferred: dict[str, int] | None = None,
    ) -> None:
        preferred = preferred or {}
        for column_index, column in enumerate(worksheet.columns, start=1):
            column_letter = get_column_letter(column_index)
            max_length = preferred.get(column_letter, min_width)
            for cell in column:
                value = cell.value
                if value is None:
                    continue
                parts = str(value).splitlines() or [""]
                max_length = max(max_length, *(len(part) for part in parts))
            worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, min_width), max_width)

    def _thin_border(self, color: str = "D8E2E8") -> Border:
        side = Side(style="thin", color=color)
        return Border(left=side, right=side, top=side, bottom=side)
