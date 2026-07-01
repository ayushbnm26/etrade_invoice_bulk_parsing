from __future__ import annotations

from io import BytesIO
import logging
from pathlib import Path
import sys
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SRC_DIR = PROJECT_ROOT / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from invoice_processing.bulk_emailer import send_bulk_admin_email  # noqa: E402
from invoice_processing.bulk_processing import BulkInvoiceProcessor  # noqa: E402
from invoice_processing.emailer import SMTPConfig, sanitize_email_exception  # noqa: E402
from invoice_processing.exceptions import EmailDeliveryError, NoInvoiceRowsError  # noqa: E402
from invoice_processing.exporters.public_transform import PUBLIC_COLUMN_LABELS  # noqa: E402
from invoice_processing.models import ParsedInvoice, PdfDocumentData, PdfPageData  # noqa: E402
from invoice_processing.parsers.invoice_parser import InvoiceParser  # noqa: E402
from invoice_processing.upload_validation import build_uploaded_file  # noqa: E402


class SequenceParser:
    def __init__(self, responses: list[ParsedInvoice | Exception]) -> None:
        self.responses = responses
        self.calls = 0

    def parse(self, _pdf_path: Path) -> ParsedInvoice:
        response = self.responses[self.calls]
        self.calls += 1
        if isinstance(response, Exception):
            raise response
        return response


def make_invoice(
    invoice_number: str = "INV-1001",
    system_ref_no: str = "SYS-1001",
    asin: str = "B000TEST",
) -> ParsedInvoice:
    item = {
        "source_file": "source.pdf",
        "document_type": "GST Invoice",
        "invoice_number": invoice_number,
        "system_ref_no": system_ref_no,
        "invoice_date": "01-Jul-2026",
        "page_no": 1,
        "si_no": 1,
        "item_description": "Test item",
        "hsn_sac": "1234",
        "asin_code": asin,
        "upc_ean": "",
        "po_no": "",
        "vendor_invoice_no": "",
        "vendor_invoice_date": "",
        "return_id": "",
        "shipment_id": "",
        "qty": 3,
        "price_per_unit": 83.33,
        "net_amount": 250.0,
        "tax_rate_raw": "9\n9",
        "tax_type_raw": "CGST SGST",
        "tax_amount_raw": "22.5\n22.5",
        "total_amount": 295.0,
        "cgst_rate": 9.0,
        "cgst_amount": 22.5,
        "sgst_rate": 9.0,
        "sgst_amount": 22.5,
        "igst_rate": None,
        "igst_amount": None,
    }
    return ParsedInvoice(
        header={
            "source_file": "source.pdf",
            "document_type": "GST Invoice",
            "page_count": 1,
            "invoice_number": invoice_number,
            "system_ref_no": system_ref_no,
        },
        summary={
            "source_file": "source.pdf",
            "invoice_number": invoice_number,
            "system_ref_no": system_ref_no,
            "total_qty_pdf": 3,
            "grand_total_pdf": 295.0,
        },
        line_items=[item],
        validation={
            "source_file": "source.pdf",
            "invoice_number": invoice_number,
            "system_ref_no": system_ref_no,
            "line_count": 1,
            "status": "PASS",
            "errors": "",
        },
    )


def make_upload(index: int, filename: str, content: bytes = b"%PDF-1.4\nfake") -> object:
    return build_uploaded_file(index, filename, content)


def workbook_from_bytes(content: bytes):
    return load_workbook(BytesIO(content))


def public_cells_as_text(workbook) -> str:
    values: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    values.append(str(cell.value))
    return "\n".join(values)


class BulkProcessingTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        logging.disable(logging.CRITICAL)

    @classmethod
    def tearDownClass(cls) -> None:
        logging.disable(logging.NOTSET)

    def process(self, uploads, parser: SequenceParser):
        with tempfile.TemporaryDirectory() as temp_dir:
            return BulkInvoiceProcessor(parser=parser).process(uploads, work_dir=Path(temp_dir))

    def test_system_ref_no_stops_at_next_header_label(self) -> None:
        cases = [
            ("System Ref No : 42000104035 Credit Note Date : 02-Jun-2026", "42000104035"),
            ("System Ref No : 42000103534 Credit Note Date : 02-Jun-2026", "42000103534"),
            ("System Ref No : 30001158056 Invoice Date : 21-May-2026", "30001158056"),
        ]

        parser = InvoiceParser()
        for text, expected in cases:
            with self.subTest(text=text):
                document = PdfDocumentData(
                    path=Path("test.pdf"),
                    page_count=1,
                    pages=[PdfPageData(page_number=1, text=text, words=[], tables=[])],
                )

                self.assertEqual(parser._parse_header(document)["system_ref_no"], expected)

    def test_two_successful_uploads_create_two_invoice_sheets(self) -> None:
        result = self.process(
            [make_upload(1, "first.pdf"), make_upload(2, "second.pdf")],
            SequenceParser([make_invoice("INV-1", "SYS-1"), make_invoice("INV-2", "SYS-2")]),
        )

        workbook = workbook_from_bytes(result.public_workbook_bytes)
        invoice_sheets = [name for name in workbook.sheetnames if name != "Processing Summary"]

        self.assertEqual(invoice_sheets, ["SYS-1", "SYS-2"])
        self.assertEqual(result.successful_invoices, 2)
        self.assertEqual(result.total_item_rows, 2)

    def test_successful_sheet_has_exact_public_columns_and_invoice_refs_per_row(self) -> None:
        result = self.process(
            [make_upload(1, "invoice.pdf")],
            SequenceParser([make_invoice("INV-777", "SYS-777")]),
        )

        worksheet = workbook_from_bytes(result.public_workbook_bytes)["SYS-777"]
        headers = [worksheet.cell(row=3, column=column).value for column in range(1, 17)]

        self.assertEqual(headers, PUBLIC_COLUMN_LABELS)
        self.assertEqual(worksheet["A4"].value, "INV-777")
        self.assertEqual(worksheet["B4"].value, "SYS-777")
        self.assertEqual(worksheet.freeze_panes, "A4")
        self.assertFalse(worksheet.sheet_view.showGridLines)

    def test_duplicate_sheet_names_are_unique(self) -> None:
        result = self.process(
            [make_upload(1, "one.pdf"), make_upload(2, "two.pdf")],
            SequenceParser([make_invoice("INV-1", "DUP"), make_invoice("INV-2", "DUP")]),
        )

        workbook = workbook_from_bytes(result.public_workbook_bytes)

        self.assertIn("DUP", workbook.sheetnames)
        self.assertIn("DUP_2", workbook.sheetnames)

    def test_long_sheet_names_are_trimmed_to_excel_limit(self) -> None:
        long_ref = "SYS-" + ("1234567890" * 5)
        result = self.process([make_upload(1, "long.pdf")], SequenceParser([make_invoice("INV", long_ref)]))

        invoice_sheet = [name for name in workbook_from_bytes(result.public_workbook_bytes).sheetnames if name != "Processing Summary"][0]

        self.assertLessEqual(len(invoice_sheet), 31)
        self.assertTrue(invoice_sheet.startswith("SYS-"))

    def test_invalid_sheet_name_characters_are_replaced(self) -> None:
        result = self.process(
            [make_upload(1, "invalid.pdf")],
            SequenceParser([make_invoice("INV", "SYS/ABC*DEF?[1]:X")]),
        )

        invoice_sheet = [name for name in workbook_from_bytes(result.public_workbook_bytes).sheetnames if name != "Processing Summary"][0]

        for invalid_char in "[]:*?/\\":
            self.assertNotIn(invalid_char, invoice_sheet)

    def test_one_invoice_failure_does_not_stop_successful_invoice(self) -> None:
        result = self.process(
            [make_upload(1, "ok.pdf"), make_upload(2, "bad.pdf")],
            SequenceParser([make_invoice("INV-OK", "SYS-OK"), NoInvoiceRowsError("No invoice table rows found")]),
        )

        workbook = workbook_from_bytes(result.public_workbook_bytes)

        self.assertEqual(result.successful_invoices, 1)
        self.assertEqual(result.failed_invoices, 1)
        self.assertEqual(workbook["bad"]["A1"].value, "Invoice Processing Failed")
        self.assertEqual(workbook["SYS-OK"]["A4"].value, "INV-OK")

    def test_all_failed_run_still_creates_workbook(self) -> None:
        result = self.process(
            [make_upload(1, "bad-one.pdf", b"not pdf"), make_upload(2, "bad-two.pdf", b"also not pdf")],
            SequenceParser([]),
        )

        workbook = workbook_from_bytes(result.public_workbook_bytes)

        self.assertEqual(result.successful_invoices, 0)
        self.assertEqual(result.failed_invoices, 2)
        self.assertEqual(workbook["bad-one"]["A1"].value, "Invoice Processing Failed")
        self.assertEqual(workbook["bad-two"]["A1"].value, "Invoice Processing Failed")

    def test_invalid_non_pdf_upload_is_per_file_failure(self) -> None:
        result = self.process([make_upload(1, "notes.txt", b"hello")], SequenceParser([]))

        self.assertEqual(result.failed_invoices, 1)
        self.assertEqual(result.outcomes[0].error.error_type, "UploadValidationError")
        self.assertEqual(result.outcomes[0].error.likely_reason, "not a PDF")

    def test_empty_pdf_is_per_file_failure(self) -> None:
        result = self.process([make_upload(1, "empty.pdf", b"")], SequenceParser([]))

        self.assertEqual(result.failed_invoices, 1)
        self.assertEqual(result.outcomes[0].error.error_type, "EmptyFileError")
        self.assertEqual(result.outcomes[0].error.likely_reason, "empty PDF")

    def test_admin_email_failure_does_not_block_public_workbook_bytes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            result = BulkInvoiceProcessor(parser=SequenceParser([make_invoice()])).process(
                [make_upload(1, "invoice.pdf")],
                work_dir=Path(temp_dir),
            )
            smtp_config = SMTPConfig(
                host="smtp.example.com",
                port=587,
                username="sender@example.com",
                app_password="secret-password",
                from_email="sender@example.com",
            )

            with patch(
                "invoice_processing.bulk_emailer.send_email",
                side_effect=EmailDeliveryError("SMTP email delivery failed: rejected"),
            ):
                email_result = send_bulk_admin_email(
                    run_result=result,
                    smtp_config=smtp_config,
                    admin_email="admin@example.com",
                )

        self.assertEqual(email_result.status, "failed")
        self.assertGreater(len(result.public_workbook_bytes), 0)

    def test_missing_smtp_config_skips_email_without_processing_failure(self) -> None:
        result = self.process([make_upload(1, "invoice.pdf")], SequenceParser([make_invoice()]))

        email_result = send_bulk_admin_email(
            run_result=result,
            smtp_config=None,
            admin_email=None,
            config_error="SMTP secrets are not configured.",
        )

        self.assertEqual(email_result.status, "skipped")
        self.assertGreater(len(result.public_workbook_bytes), 0)

    def test_smtp_error_sanitizer_redacts_credentials(self) -> None:
        smtp_config = SMTPConfig(
            host="smtp.example.com",
            port=587,
            username="sender@example.com",
            app_password="secret-password",
            from_email="sender@example.com",
        )

        message = sanitize_email_exception(
            RuntimeError("login failed for sender@example.com with secret-password"),
            smtp_config,
        )

        self.assertNotIn("sender@example.com", message)
        self.assertNotIn("secret-password", message)
        self.assertIn("[redacted]", message)

    def test_internal_workbook_keeps_exception_details_not_public_workbook(self) -> None:
        raw_path_error = RuntimeError(r"boom at C:\Users\ayush\Desktop\secret\parser.py")
        result = self.process([make_upload(1, "broken.pdf")], SequenceParser([raw_path_error]))

        public_workbook = workbook_from_bytes(result.public_workbook_bytes)
        internal_workbook = workbook_from_bytes(result.internal_workbook_bytes)
        public_text = public_cells_as_text(public_workbook)
        internal_text = public_cells_as_text(internal_workbook)

        self.assertIn(r"C:\Users\ayush\Desktop\secret\parser.py", internal_text)
        self.assertIn("Traceback", internal_text)
        self.assertNotIn("Traceback", public_text)
        self.assertNotIn(r"C:\Users", public_text)
        self.assertNotIn("secret", public_text)


if __name__ == "__main__":
    unittest.main()
